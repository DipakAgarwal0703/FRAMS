from sys import path
import csv
from tkinter import*
from tkinter import ttk
from PIL import Image,ImageTk
import os
import mysql.connector
import cv2
from tkinter import filedialog as fd
import numpy as np
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import date
import datetime
from datetime import datetime
from tkinter import messagebox
import time
import pkg_resources
import face_recognition
#haar=pkg_resources.resource_filename('cv2','data/haarcascade_frontalface_default.xml')
directory="C:\\Users\\DIPAK\\OneDrive\\Desktop\\FINAL FACE\\"
# from PIL import ImageGrab

class Attendance_Section:
    def __init__(self,root):
        #Setting Up window form
        self.root=root
        self.root.geometry("670x230+480+390")
        self.root.title("Mark Attendance")      
        self.root.resizable(False,False)
        self.root.focus_force()
   
        bg_img = Label(self.root,bg="white")
        bg_img.place(x=0,y=0,relwidth=1,relheight=1)  

        # Variable declaration
        self.var_subject=StringVar()
        
        left_frame = LabelFrame(bg_img,bd=2,bg="white",relief=RIDGE,text="Attendance Sheets",font=("verdana",12,"bold"),fg="black")
        left_frame.place(x=185,y=0,width=300,height=80)

        # Subject Selection
        sub_combo=ttk.Combobox(left_frame,textvariable=self.var_subject,width=15,font=("verdana",12,"bold"),state="readonly")
        sub_combo["values"]=("Select Course","Computer Architecture","Design Analysis of Algorithm","Formal Language & Automata Theory","Discrete Mathematics","Biology","Environmental Science")
        sub_combo.current(0)
        sub_combo.place(x=5,y=15,width=280,height=25)
        

        # Create buttons below the section 
        # Take Attendance Button
        start_atten=Image.open(r"Images\Start_Attendance.png")     
        start_atten=start_atten.resize((300,50),Image.ANTIALIAS)
        self.photostart_atten=ImageTk.PhotoImage(start_atten) 

        b1=Button(bg_img,image=self.photostart_atten,command=self.facerecognize,cursor="hand2",bd=2,relief=RIDGE,fg="white",bg="black",activeforeground="black",activebackground="black")
        b1.place(x=5,y=120,width=300,height=50)

        # Save Attendance Button
        save_atten=Image.open(r"Images\Save_Attendance.png")     
        save_atten=save_atten.resize((300,50),Image.ANTIALIAS)
        self.photosave_atten=ImageTk.PhotoImage(save_atten) 

        b2=Button(bg_img,image=self.photosave_atten,command=self.txttolist,cursor="hand2",bd=2,relief=RIDGE,fg="white",bg="black",activeforeground="black",activebackground="black")
        b2.place(x=355,y=120,width=300,height=50)
        
    # ==========Mark Attendance===========
    def browse(self):
        filename = fd.askopenfilename()
        return filename

    def save_raw_Attendance(self,i,r,n,d):
        today = str(date.today())
        txt="raw_"+today+".txt"
        filepath = directory
        file=filepath+txt
        with open(filepath+"raw_"+today+".txt", "a+",newline="\n") as fp:
                myDatalist=fp.readlines()
                name_list=[]
                for line in myDatalist:
                    entry=line.split((","))
                    name_list.append(entry[0])
                if((n not in name_list)):
                    now=datetime.now()
                    d1=now.strftime("%d/%m/%Y")
                    dtString=now.strftime("%H:%M:%S")
                    fp.writelines(f"\n{n}")

        def remove_repeated(file):
            openFile = open(file, "r") 
            writeFile = open("text_student.txt", "w") #bug
            
            #Store traversed lines
            tmp = set() 
            for txtLine in openFile: 
            
            #Check new line
                if txtLine not in tmp: 
                    writeFile.write(txtLine) 
            
            #Add new traversed line to tmp 
                    tmp.add(txtLine)         
            openFile.close() 
            writeFile.close()
        remove_repeated(file)
    

    def markA(self,namelist):
        sub=self.var_subject.get()
        if sub!="Select Course":

            filename = fd.askopenfilename()
            column=0
            today = date.today()
            td=str(today)
            datedata=""
            datad=0
            column=0
            td=str(today)
            wb=load_workbook(filename)
            ws=wb.active
            sheets = wb.sheetnames
            mont=str(today.month)
            if mont=="9":
                ws = wb[sheets[0]]
            elif mont=="10":
                ws = wb[sheets[1]]
                print(ws)
            elif mont=="11":
                ws = wb[sheets[2]]
            elif mont=="12":
                ws = wb[sheets[3]]

            dateindex=0
            for dtrows in range(1,2):
                for dtcol in range(3,12):
                    char=get_column_letter(dtcol)
                    val=ws[char+str(dtrows)].value
                    val=str(val)
                    datedata=val[0:10]
                    if td==datedata:
                        dateindex=dtcol
                        break
            if td==datedata:
                column=dateindex
            else:
                messagebox.showerror("Hold on","No class today!!",parent=self.root)
            rowdata=0
            for rows in range(3,7):
                for col in range(2,3):
                    char=get_column_letter(col)
                    val=ws[char+str(rows)].value
                    for i in namelist:
                        if i==val:
                            row=rowdata+rows
                            c1 = ws.cell(row, column)
                            c1.value = "P" 
            
            wb.save(r"C:\Users\DIPAK\OneDrive\Desktop\FACE2\Attendance_Sheet\sheet_"+sub+".xlsx")
            messagebox.showinfo("Saved","Attendance Successfully saved!",parent=self.root)
        else:
            messagebox.showerror("Error","Please Select the subject first!",parent=self.root)
        
    def txttolist(self):
        my_file = open("text_student.txt", "r")
        print (my_file.readline())
        
        # reading the file
        data = my_file.read()
        
        # replacing end splitting the text  when newline ('\n') is seen.
        data_into_list = data.split("\n")
        correctlist=data_into_list.pop()
        my_file.close()
        self.markA(data_into_list)
        

    # ********* face recognition ******************************
    # def facerecognize(self):
    #     def draw_boundray(img,classifier,scaleFactor,minNeighbours,color,text,clf):
    #         gray_image1=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)    
    #         features=classifier.detectMultiScale(gray_image1,scaleFactor,minNeighbours)

    #         coord=[]
        
    #         for (x,y,w,h) in features:
    #             cv2.rectangle(img,(x,y),(x+w,y+h),(0,255,0),3)
    #             id,predict=clf.predict(gray_image1[y:y+h,x:x+w])
    #             print (id)
    #             confidence=int((100*(1-predict/300)))
                
    #             conn=mysql.connector.connect(host="localhost",username="root",password="dipak0703@AG",database="face_recognition")
    #             my_cursor=conn.cursor()
                
    #              #As much as we want
    #             my_cursor.execute("select Student_Name from student where Student_ID="+str(id))
    #             n=my_cursor.fetchone()
    #             if n:
    #                 n="+".join(n)

    #             if confidence>77:
    #                 cv2.putText(img,f"Name:{n}",(x,y-30),cv2.FONT_HERSHEY_COMPLEX,0.8,(255,255,255),3)
    #                 self.save_raw_Attendance(n)

    #             else:
    #                 cv2.rectangle(img,(x,y),(x+w,y+h),(0,0,255),3)
    #                 cv2.putText(img,"Unknown Face",(x,y-5),cv2.FONT_HERSHEY_COMPLEX,0.8,(255,255,255),3)
                       
    #             coord=[x,y,w,y]
                
    #         return coord
        
    #     def recognize(img,clf,faceCascade):
    #         coord=draw_boundray(img,faceCascade,1.1,10,(255,25,255),"Face",clf) 
    #         return img
        
    #     faceCascade=cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
    #     clf=cv2.face.LBPHFaceRecognizer_create()
    #     clf.read("classifier.xml") 
        
    #     video_cap=cv2.VideoCapture(0)

    #     while (True):
    #         ret,img=video_cap.read()
    #         img=recognize(img,clf,faceCascade)
    #         cv2.imshow("WELCOME TO FACE RECOGNITION",img)
            
    #         if cv2.waitKey(1)==13:
    #             break
    #     video_cap.release()
    #     cv2.destroyAllWindows()
 

    def facerecognize(self):
        def draw_boundray(img,classifier,scaleFactor,minNeighbours,color,text,clf):
            gray_image1=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)    
            features=classifier.detectMultiScale(gray_image1,scaleFactor,minNeighbours)
            
            coord=[]
        
            for (x,y,w,h) in features:
                cv2.rectangle(img,(x,y),(x+w,y+h),(0,255,0),3)
                im,predict=clf.predict(gray_image1[y:y+h,x:x+w])
                print (im)
                confidence=int((100*(1-predict/300)))
                
                conn=mysql.connector.connect(host="localhost",username="root",password="dipak0703@AG",database="face_recognition")
                my_cursor=conn.cursor()
                
                 #As much as we want
                my_cursor.execute("select Student_Name from student where Student_id="+str(im))
                n=my_cursor.fetchone()
                if n:
                    n="+".join(n)
                
                my_cursor.execute("select Roll_No from student where Student_id="+str(im))
                r=my_cursor.fetchone()
                if r:    
                    r="+".join(r)
                
                my_cursor.execute("select Department from student where Student_id="+str(im))
                d=my_cursor.fetchone()
                if d:
                    d="+".join(d)
                    
                my_cursor.execute("select Student_ID from student where Student_id="+str(im))
                i=my_cursor.fetchone()
                if i:
                    i="+".join(i)
                
                
                
                if confidence>77:
                    cv2.putText(img,f"ID:{i}",(x,y-75),cv2.FONT_HERSHEY_COMPLEX,0.8,(255,255,255),3)
                    cv2.putText(img,f"Roll:{r}",(x,y-55),cv2.FONT_HERSHEY_COMPLEX,0.8,(255,255,255),3)
                    cv2.putText(img,f"Name:{n}",(x,y-30),cv2.FONT_HERSHEY_COMPLEX,0.8,(255,255,255),3)
                    cv2.putText(img,f"Department:{d}",(x,y-5),cv2.FONT_HERSHEY_COMPLEX,0.8,(255,255,255),3)
                    # self.mark_attendance(i,r,n,d)
                    self.save_raw_Attendance(i,r,n,d)
                    
                else:
                    cv2.rectangle(img,(x,y),(x+w,y+h),(0,0,255),3)
                    cv2.putText(img,"Unknow Face",(x,y-5),cv2.FONT_HERSHEY_COMPLEX,0.8,(255,255,255),3)
                       
                coord=[x,y,w,y]
                
            return coord
        
        def recognize(img,clf,faceCascade):
            coord=draw_boundray(img,faceCascade,1.1,10,"(255,25,255)","Face",clf) 
            return img
        
        faceCascade=cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
        clf=cv2.face.LBPHFaceRecognizer_create()
        clf.read("classifier.xml") 
        
        video_cap=cv2.VideoCapture(0)
        
        while (True):
            ret,img=video_cap.read()
            img=recognize(img,clf,faceCascade)
            cv2.imshow("WELCOME TO FACE RECOGNITION",img)
            
            if cv2.waitKey(1)==13:
                break
        video_cap.release()
        cv2.destroyAllWindows()

        
if __name__ == "__main__":
    root=Tk()
    obj=Attendance_Section(root)
    root.mainloop()